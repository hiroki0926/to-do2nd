import openpyxl
import firebase_admin
from firebase_admin import credentials, db

# Firebaseの資格情報を設定
cred = credentials.Certificate("todo-app-6f566-firebase-adminsdk-f7fm0-89b3e811c8.json")
firebase_admin.initialize_app(cred, {"databaseURL": "https://todo-app-6f566-default-rtdb.firebaseio.com"})
# Firebaseからデータを取得
ref = db.reference("/tasks")
result = ref.get()

# 新しいExcelワークブックとシートを作成
workbook = openpyxl.Workbook()
sheet = workbook.active

# ヘッダーを追加
sheet['A1'] = 'タスクID'
sheet['B1'] = 'タスクテキスト'
sheet['C1'] = 'カテゴリ'
sheet['D1'] = '完了済み'

# データを埋める
row = 2
for task_id, task_data in result.items():
    sheet.cell(row=row, column=1, value=task_id)
    sheet.cell(row=row, column=2, value=task_data.get('text', ''))
    sheet.cell(row=row, column=3, value=task_data.get('category', ''))
    sheet.cell(row=row, column=4, value=task_data.get('completed', False))
    row += 1

# Excelファイルを保存
workbook.save('tasks.xlsx')
